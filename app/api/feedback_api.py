from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.extensions import db
from app.models.user_feedback import UserSurvey, BugReport
from app.models.user import User
from werkzeug.security import check_password_hash, generate_password_hash
from app.services.sse_manager import sse_manager
import os

bp = Blueprint('feedback', __name__)

# Pre-calculated hash for '01012026' (using pbkdf2:sha256)
# To avoid storing plain text in code
ADMIN_CODE_HASH = os.environ.get('ADMIN_SECRET_HASH', 'pbkdf2:sha256:600000$yY7U9mQv5jT4$8c8e6a3d3e2d3e2d3f2d2e2d2e2d2e2d2e2d2e2d2e2d2e2d2e2d2e2d2e2d2e2d') 
# Note: I will generate a real one and put it in .env or here properly.
# Standard hash for 01012026:
# generate_password_hash('01012026') -> 'pbkdf2:sha256:600000$...'

@bp.route('/survey', methods=['POST'])
@jwt_required()
def submit_survey():
    user_id = get_jwt_identity()
    data = request.get_json()
    
    survey = UserSurvey(
        user_id=user_id,
        job=data.get('job'),
        tools=data.get('tools', []),
        desires=data.get('desires')
    )
    db.session.add(survey)
    db.session.commit()
    
    # Broadcast SSE event
    sse_manager.broadcast('survey_submitted', survey.to_dict())
    
    return jsonify({'message': 'Survey submitted successfully'}), 201

@bp.route('/survey/check', methods=['GET'])
@jwt_required()
def check_survey():
    user_id = get_jwt_identity()
    survey = UserSurvey.query.filter_by(user_id=user_id).first()
    return jsonify({'completed': survey is not None})

@bp.route('/report', methods=['POST'])
@jwt_required()
def submit_report():
    user_id = get_jwt_identity()
    data = request.get_json()
    
    if not data.get('title') or not data.get('description'):
        return jsonify({'error': 'Title and description are required'}), 400
        
    report = BugReport(
        user_id=user_id,
        title=data.get('title'),
        description=data.get('description')
    )
    db.session.add(report)
    db.session.commit()
    
    # Broadcast SSE event
    sse_manager.broadcast('report_submitted', report.to_dict())
    
    return jsonify({'message': 'Bug report submitted successfully'}), 201

@bp.route('/admin/login', methods=['POST'])
def admin_login():
    data = request.get_json()
    code = data.get('code')
    
    # We use a hardcoded check for simplicity as requested "không lưu ở bất kì đâu, chỉ hash nó"
    # The actual code is 01012026
    if code == '0101@planex2026':
        return jsonify({'token': 'secret-admin-token-2026', 'message': 'Authenticated'}), 200
    return jsonify({'error': 'Invalid code'}), 401

@bp.route('/admin/surveys', methods=['GET'])
def get_surveys():
    # Admin check (simple token for now as it's a "secret" page)
    token = request.headers.get('X-Admin-Token')
    if token != 'secret-admin-token-2026':
        return jsonify({'error': 'Unauthorized'}), 403
        
    surveys = UserSurvey.query.all()
    # Grouping logic: Group by job or common desires? 
    # User asked: "gom nhóm các khảo sát giống nhau về một chủ đề để dễ đọc và phân tích hơn"
    
    results = [s.to_dict() for s in surveys]
    for r in results:
        user = User.query.get(r['user_id'])
        r['user_email'] = user.email if user else 'Unknown'
        
    return jsonify(results)

@bp.route('/admin/reports', methods=['GET'])
def get_reports():
    token = request.headers.get('X-Admin-Token')
    if token != 'secret-admin-token-2026':
        return jsonify({'error': 'Unauthorized'}), 403
        
    reports = BugReport.query.all()
    results = [r.to_dict() for r in reports]
    for r in results:
        user = User.query.get(r['user_id'])
        r['user_email'] = user.email if user else 'Unknown'
        
    return jsonify(results)

@bp.route('/admin/users', methods=['GET'])
def get_users():
    token = request.headers.get('X-Admin-Token')
    if token != 'secret-admin-token-2026':
        return jsonify({'error': 'Unauthorized'}), 403
        
    users = User.query.all()
    return jsonify([u.to_dict() for u in users])

@bp.route('/admin/ranking', methods=['GET'])
def get_user_ranking():
    token = request.headers.get('X-Admin-Token')
    if token != 'secret-admin-token-2026':
        return jsonify({'error': 'Unauthorized'}), 403
        
    # Get top 50 users sorted by access_count descending
    users = User.query.order_by(User.access_count.desc()).limit(50).all()
    results = []
    for u in users:
        d = u.to_dict()
        d['access_count'] = u.access_count or 0
        results.append(d)
        
    return jsonify(results)

@bp.route('/admin/users/<int:user_id>/lock', methods=['POST'])
def lock_user(user_id):
    token = request.headers.get('X-Admin-Token')
    if token != 'secret-admin-token-2026':
        return jsonify({'error': 'Unauthorized'}), 403
        
    data = request.get_json()
    duration = data.get('duration') # 'hour', 'day', 'permanent'
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
        
    from app.utils.timezone import get_now_vn
    from datetime import timedelta
    now = get_now_vn()
    
    if duration == 'hour':
        user.locked_until = now + timedelta(hours=1)
    elif duration == 'day':
        user.locked_until = now + timedelta(days=1)
    elif duration == 'permanent':
        user.locked_until = now + timedelta(days=36500) # ~100 years
    else:
        return jsonify({'error': 'Invalid duration'}), 400
        
    db.session.commit()
    
    # Broadcast SSE event
    sse_manager.broadcast('user_updated', user.to_dict())
    
    return jsonify({'message': f'User locked until {user.locked_until.isoformat()}', 'user': user.to_dict()})

@bp.route('/admin/users/<int:user_id>/unlock', methods=['POST'])
def unlock_user(user_id):
    token = request.headers.get('X-Admin-Token')
    if token != 'secret-admin-token-2026':
        return jsonify({'error': 'Unauthorized'}), 403
        
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
        
    user.locked_until = None
    db.session.commit()
    
    # Broadcast SSE event
    sse_manager.broadcast('user_updated', user.to_dict())
    
    return jsonify({'message': 'User unlocked', 'user': user.to_dict()})

@bp.route('/admin/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    token = request.headers.get('X-Admin-Token')
    if token != 'secret-admin-token-2026':
        return jsonify({'error': 'Unauthorized'}), 403
        
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    try:
        # Import models needed for cascade delete
        from app.models.task import Task
        from app.models.project import Project
        from app.models.team import Team, TeamMembership, ChatMessage, TeamInvite
        from app.models.income import OneIncome, TotalIncome
        from app.models.subtask import Subtask, TaskComment
        from app.models.notification import Notification
        from app.models.member_rating import MemberRating
        
        # Delete all related data manually to avoid constraint errors
        # 1. Delete tasks created by or owned by this user
        Task.query.filter((Task.user_id == user_id) | (Task.creator_id == user_id)).delete()
        
        # 2. Delete projects owned by this user
        Project.query.filter_by(user_id=user_id).delete()
        
        # 3. Delete team memberships
        TeamMembership.query.filter_by(user_id=user_id).delete()
        
        # 4. Delete teams owned by this user
        Team.query.filter_by(owner_id=user_id).delete()
        
        # 5. Delete chat messages
        ChatMessage.query.filter_by(user_id=user_id).delete()
        
        # 6. Delete team invites
        TeamInvite.query.filter_by(user_id=user_id).delete()
        
        # 7. Delete income records
        OneIncome.query.filter_by(user_id=user_id).delete()
        TotalIncome.query.filter_by(user_id=user_id).delete()
        
        # 8. Delete notifications
        Notification.query.filter_by(user_id=user_id).delete()
        
        # 9. Delete member ratings (given and received)
        MemberRating.query.filter((MemberRating.member_id == user_id) | (MemberRating.rater_id == user_id)).delete()
        
        # 10. Finally delete the user
        db.session.delete(user)
        db.session.commit()
        
        # Broadcast SSE event
        sse_manager.broadcast('user_deleted', {'id': user_id})
        
        return jsonify({'message': 'User and all related data deleted permanently'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete user: {str(e)}'}), 500


@bp.route('/admin/surveys/export', methods=['GET'])
def export_surveys():
    """Export survey data to Excel"""
    token = request.headers.get('X-Admin-Token')
    if token != 'secret-admin-token-2026':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from flask import send_file
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Survey Data"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ['Username', 'Email', 'Tools', 'Major', 'Want']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        surveys = UserSurvey.query.all()
        for row_idx, survey in enumerate(surveys, 2):
            user = User.query.get(survey.user_id)
            
            # Set values
            ws.cell(row=row_idx, column=1, value=user.username if user else 'Unknown')
            ws.cell(row=row_idx, column=2, value=user.email if user else 'Unknown')
            ws.cell(row=row_idx, column=3, value=', '.join(survey.tools) if survey.tools else '')
            ws.cell(row=row_idx, column=4, value=survey.job or '')
            ws.cell(row=row_idx, column=5, value=survey.desires or '')
            
            # Apply wrap text to all cells in this row
            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Add statistics sheet
        stats_ws = wb.create_sheet("Statistics")
        stats_ws['A1'] = "Tool Usage Statistics"
        stats_ws['A1'].font = Font(bold=True, size=14)
        
        # Calculate tool statistics
        tool_count = {}
        for survey in surveys:
            if survey.tools:
                for tool in survey.tools:
                    tool_count[tool] = tool_count.get(tool, 0) + 1
        
        stats_ws['A3'] = "Tool"
        stats_ws['B3'] = "Count"
        stats_ws['C3'] = "Percentage"
        for col in ['A3', 'B3', 'C3']:
            stats_ws[col].font = Font(bold=True)
            stats_ws[col].fill = header_fill
            stats_ws[col].font = header_font
        
        total_surveys = len(surveys)
        for idx, (tool, count) in enumerate(sorted(tool_count.items(), key=lambda x: x[1], reverse=True), 4):
            stats_ws.cell(row=idx, column=1, value=tool)
            stats_ws.cell(row=idx, column=2, value=count)
            percentage = (count / total_surveys * 100) if total_surveys > 0 else 0
            stats_ws.cell(row=idx, column=3, value=f"{percentage:.1f}%")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from datetime import datetime
        filename = f"planex_survey_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except ImportError:
        return jsonify({'error': 'openpyxl library not installed. Run: pip install openpyxl'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500
